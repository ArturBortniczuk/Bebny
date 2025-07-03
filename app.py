from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash
import openpyxl
import math
import os
import json
import logging
import hashlib
import secrets
from datetime import datetime, timedelta
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
app.permanent_session_lifetime = timedelta(hours=8)  # Sesja na 8 godzin

# Konfiguracja logowania
logging.basicConfig(level=logging.INFO)

# Cache dla danych Excel
excel_data_cache = {}

# Pobierz ścieżkę do katalogu, w którym znajduje się plik app.py
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(BASE_DIR, 'wszystkiekable.xlsx')

# Ścieżka do pliku z danymi klientów i bębnów
CLIENTS_DATA_FILE = os.path.join(BASE_DIR, 'clients_data.json')
PASSWORDS_FILE = os.path.join(BASE_DIR, 'client_passwords.json')

def init_mock_data():
    """Inicializuje mock data jeśli nie istnieją"""
    
    # Mock data dla klientów i bębnów
    mock_clients_data = [
        {
            "KOD_BEBNA": "BEB001",
            "NAZWA": "Bęben stalowy 200cm",
            "CECHA": "Standardowy",
            "DATA_ZWROTU_DO_DOSTAWCY": "2024-12-15",
            "KON_DOSTAWCA": "DOSTAWCA_A",
            "PELNA_NAZWA_KONTRAHENTA": "ELEKTRO-BUD Sp. z o.o.",
            "NIP": "1234567890",
            "TYP_DOK": "WZ",
            "NR_DOKUMENTUPZ": "WZ/2024/001",
            "Data_przyjęcia_na_stan": "2024-01-15",
            "KONTRAHENT": "ELEKTRO-BUD",
            "STATUS": "Wypożyczony",
            "DATA_WYDANIA": "2024-01-15"
        },
        {
            "KOD_BEBNA": "BEB002",
            "NAZWA": "Bęben stalowy 180cm",
            "CECHA": "Wzmocniony",
            "DATA_ZWROTU_DO_DOSTAWCY": "2024-11-20",
            "KON_DOSTAWCA": "DOSTAWCA_B",
            "PELNA_NAZWA_KONTRAHENTA": "ELEKTRO-BUD Sp. z o.o.",
            "NIP": "1234567890",
            "TYP_DOK": "WZ",
            "NR_DOKUMENTUPZ": "WZ/2024/002",
            "Data_przyjęcia_na_stan": "2024-02-01",
            "KONTRAHENT": "ELEKTRO-BUD",
            "STATUS": "Wypożyczony",
            "DATA_WYDANIA": "2024-02-01"
        },
        {
            "KOD_BEBNA": "BEB003",
            "NAZWA": "Bęben drewniany 150cm",
            "CECHA": "Ekologiczny",
            "DATA_ZWROTU_DO_DOSTAWCY": "2024-10-30",
            "KON_DOSTAWCA": "DOSTAWCA_A",
            "PELNA_NAZWA_KONTRAHENTA": "TECH-ELEKTRO S.A.",
            "NIP": "9876543210",
            "TYP_DOK": "WZ",
            "NR_DOKUMENTUPZ": "WZ/2024/003",
            "Data_przyjęcia_na_stan": "2024-03-10",
            "KONTRAHENT": "TECH-ELEKTRO",
            "STATUS": "Wypożyczony",
            "DATA_WYDANIA": "2024-03-10"
        },
        {
            "KOD_BEBNA": "BEB004",
            "NAZWA": "Bęben stalowy 220cm",
            "CECHA": "Duży",
            "DATA_ZWROTU_DO_DOSTAWCY": "2024-12-01",
            "KON_DOSTAWCA": "DOSTAWCA_C",
            "PELNA_NAZWA_KONTRAHENTA": "TECH-ELEKTRO S.A.",
            "NIP": "9876543210",
            "TYP_DOK": "WZ",
            "NR_DOKUMENTUPZ": "WZ/2024/004",
            "Data_przyjęcia_na_stan": "2024-01-20",
            "KONTRAHENT": "TECH-ELEKTRO",
            "STATUS": "Wypożyczony",
            "DATA_WYDANIA": "2024-01-20"
        }
    ]
    
    if not os.path.exists(CLIENTS_DATA_FILE):
        with open(CLIENTS_DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(mock_clients_data, f, ensure_ascii=False, indent=2)
        logging.info("Utworzono plik z danymi klientów")
    
    if not os.path.exists(PASSWORDS_FILE):
        with open(PASSWORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump({}, f, ensure_ascii=False, indent=2)
        logging.info("Utworzono plik z hasłami klientów")

def load_clients_data():
    """Ładuje dane klientów z pliku JSON"""
    try:
        with open(CLIENTS_DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        init_mock_data()
        return load_clients_data()
    except Exception as e:
        logging.error(f"Błąd ładowania danych klientów: {e}")
        return []

def load_passwords():
    """Ładuje hasła klientów z pliku JSON"""
    try:
        with open(PASSWORDS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}
    except Exception as e:
        logging.error(f"Błąd ładowania haseł: {e}")
        return {}

def save_passwords(passwords):
    """Zapisuje hasła klientów do pliku JSON"""
    try:
        with open(PASSWORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(passwords, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logging.error(f"Błąd zapisywania haseł: {e}")
        return False

def hash_password(password):
    """Haszuje hasło używając SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    """Weryfikuje hasło"""
    return hash_password(password) == hashed

def login_required(f):
    """Dekorator wymagający logowania"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'nip' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def get_client_by_nip(nip):
    """Pobiera dane klienta po NIP"""
    clients_data = load_clients_data()
    for client in clients_data:
        if client.get('NIP') == nip:
            return client
    return None

def get_client_drums(nip):
    """Pobiera bębny klienta po NIP"""
    clients_data = load_clients_data()
    client_drums = []
    
    for item in clients_data:
        if item.get('NIP') == nip:
            client_drums.append(item)
    
    return client_drums

# Istniejące funkcje kalkulatora (load_excel_data, etc.)
def load_excel_data():
    """Ładuje dane z pliku Excel z cache'owaniem (bez pandas)"""
    global excel_data_cache
    
    if not excel_data_cache:
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            # Ładowanie arkusza "Kable"
            kable_sheet = workbook['Kable']
            kable_data = []
            headers_kable = [cell.value for cell in kable_sheet[1]]
            
            for row in kable_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Sprawdź czy pierwsza kolumna nie jest pusta
                    row_dict = {headers_kable[i]: row[i] for i in range(len(headers_kable)) if i < len(row)}
                    kable_data.append(row_dict)
            
            # Ładowanie arkusza "Wymiary"
            bebny_sheet = workbook['Wymiary']
            bebny_data = []
            headers_bebny = [cell.value for cell in bebny_sheet[1]]
            
            for row in bebny_sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Sprawdź czy pierwsza kolumna nie jest pusta
                    row_dict = {headers_bebny[i]: row[i] for i in range(len(headers_bebny)) if i < len(row)}
                    bebny_data.append(row_dict)
            
            excel_data_cache['kable_data'] = kable_data
            excel_data_cache['bebny_data'] = bebny_data
            
            logging.info("Dane Excel załadowane pomyślnie")
        except Exception as e:
            logging.error(f"Błąd ładowania danych Excel: {e}")
            # Użyj pustych danych jeśli Excel nie istnieje
            excel_data_cache['kable_data'] = []
            excel_data_cache['bebny_data'] = []
    
    return excel_data_cache['kable_data'], excel_data_cache['bebny_data']

def get_kable_options():
    """Pobiera opcje kabli i przekrojów żył"""
    kable_data, _ = load_excel_data()
    typy_kabli = list(set([kabel['Nazwa'] for kabel in kable_data if kabel.get('Nazwa')]))
    opcje_kabli = {}
    
    for kabel_typ in typy_kabli:
        przekroje = list(set([
            kabel['Liczba i przekrój żył'] 
            for kabel in kable_data 
            if kabel.get('Nazwa') == kabel_typ and kabel.get('Liczba i przekrój żył')
        ]))
        opcje_kabli[kabel_typ] = przekroje
    
    return opcje_kabli

# Reszta funkcji kalkulatora (validate_input_data, calculate_cable_on_drum, etc.)
def validate_input_data(nazwa_kabla, liczba_przekroj, dlugosc_kabla):
    """Waliduje dane wejściowe"""
    errors = []
    
    if not nazwa_kabla:
        errors.append("Nie wybrano typu kabla")
    
    if not liczba_przekroj:
        errors.append("Nie wybrano przekroju żył")
    
    try:
        dlugosc = float(dlugosc_kabla)
        if dlugosc <= 0:
            errors.append("Długość kabla musi być większa od 0")
        if dlugosc > 10000:
            errors.append("Długość kabla wydaje się zbyt duża (max 10000m)")
    except (ValueError, TypeError):
        errors.append("Długość kabla musi być liczbą")
    
    return errors

def calculate_cable_on_drum(beben, srednica_kabla, dlugosc_kabla):
    """Oblicza długość kabla na bębnie z lepszą precyzją"""
    warstwa = 0
    calkowita_dlugosc = 0
    bęben_szerokosc = beben['szerokość']
    max_warstwy = 50  # Zabezpieczenie przed nieskończoną pętlą
    
    while calkowita_dlugosc < dlugosc_kabla and warstwa < max_warstwy:
        aktualna_średnica_warstwy = beben['średnica wewnętrzna'] + warstwa * srednica_kabla * 2
        
        if aktualna_średnica_warstwy > beben['Średnica'] - 5:  # Zmniejszony margines
            break
        
        obwod_warstwy = math.pi * aktualna_średnica_warstwy
        liczba_zwojów_na_warstwie = math.floor(bęben_szerokosc / srednica_kabla)
        
        długość_na_warstwie = liczba_zwojów_na_warstwie * obwod_warstwy / 100
        calkowita_dlugosc += długość_na_warstwie
        warstwa += 1
    
    return calkowita_dlugosc, warstwa

def find_suitable_drums(srednica_kabla, promien_giecia, dlugosc_kabla, bebny_data, masa_kabla_na_km):
    """Znajduje wszystkie odpowiednie bębny z dodatkowymi informacjami"""
    minimalna_wewnetrzna = promien_giecia * 2
    odpowiednie_bebny = []
    
    for beben in bebny_data:
        if beben.get('średnica wewnętrzna', 0) >= minimalna_wewnetrzna:
            calkowita_dlugosc, liczba_warstw = calculate_cable_on_drum(beben, srednica_kabla, dlugosc_kabla)
            
            if calkowita_dlugosc >= dlugosc_kabla:
                masa_kabla = (dlugosc_kabla / 1000) * masa_kabla_na_km
                masa_bębna = beben.get('Waga', 0)
                suma_wag = masa_kabla + masa_bębna
                
                wykorzystanie_procent = (dlugosc_kabla / calkowita_dlugosc) * 100
                
                odpowiednie_bebny.append({
                    'beben': beben,
                    'masa_kabla': masa_kabla,
                    'masa_bębna': masa_bębna,
                    'suma_wag': suma_wag,
                    'wykorzystanie_procent': wykorzystanie_procent,
                    'liczba_warstw': liczba_warstw,
                    'max_dlugosc': calkowita_dlugosc
                })
    
    return sorted(odpowiednie_bebny, key=lambda x: x['suma_wag'])

# Routes

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Strona logowania"""
    if request.method == 'POST':
        nip = request.form.get('nip', '').strip()
        password = request.form.get('password', '').strip()
        
        if not nip:
            flash('Proszę wprowadzić NIP', 'error')
            return render_template('login.html')
        
        # Sprawdź czy klient istnieje
        client = get_client_by_nip(nip)
        if not client:
            flash('Nie znaleziono klienta o podanym NIP', 'error')
            return render_template('login.html')
        
        passwords = load_passwords()
        
        # Sprawdź czy hasło jest już ustawione
        if nip not in passwords:
            # Pierwsze logowanie - pozwól ustawić hasło
            if not password:
                flash('Pierwsze logowanie - ustaw hasło dla swojego konta', 'info')
                return render_template('login.html', nip=nip, first_login=True)
            
            # Zapisz nowe hasło
            passwords[nip] = hash_password(password)
            if save_passwords(passwords):
                session.permanent = True
                session['nip'] = nip
                session['company_name'] = client['PELNA_NAZWA_KONTRAHENTA']
                flash('Hasło zostało ustawione. Zostałeś zalogowany.', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Błąd podczas zapisywania hasła', 'error')
                return render_template('login.html')
        
        # Logowanie z istniejącym hasłem
        if not password:
            flash('Proszę wprowadzić hasło', 'error')
            return render_template('login.html', nip=nip)
        
        if verify_password(password, passwords[nip]):
            session.permanent = True
            session['nip'] = nip
            session['company_name'] = client['PELNA_NAZWA_KONTRAHENTA']
            flash('Zalogowano pomyślnie', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Nieprawidłowe hasło', 'error')
            return render_template('login.html', nip=nip)
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Wylogowanie"""
    session.clear()
    flash('Zostałeś wylogowany', 'info')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Panel główny klienta"""
    nip = session['nip']
    company_name = session['company_name']
    
    # Pobierz bębny klienta
    drums = get_client_drums(nip)
    
    # Statystyki
    total_drums = len(drums)
    drums_with_return_date = len([d for d in drums if d.get('DATA_ZWROTU_DO_DOSTAWCY')])
    
    return render_template('dashboard.html', 
                         drums=drums,
                         company_name=company_name,
                         total_drums=total_drums,
                         drums_with_return_date=drums_with_return_date)

@app.route('/my-drums')
@login_required
def my_drums():
    """Strona z bębnami klienta"""
    nip = session['nip']
    drums = get_client_drums(nip)
    
    # Sortuj po dacie zwrotu
    drums_sorted = sorted(drums, 
                         key=lambda x: x.get('DATA_ZWROTU_DO_DOSTAWCY', '9999-12-31'))
    
    return render_template('my_drums.html', drums=drums_sorted)

@app.route('/calculator')
@login_required
def calculator():
    """Kalkulator bębnów (dla zalogowanych użytkowników)"""
    try:
        opcje_kabli = get_kable_options()
        return render_template('calculator.html', opcje_kabli=opcje_kabli)
    except Exception as e:
        logging.error(f"Błąd na stronie kalkulatora: {e}")
        flash('Błąd ładowania kalkulatora', 'error')
        return redirect(url_for('dashboard'))

@app.route('/')
def index():
    """Strona główna - przekierowanie do logowania lub dashboard"""
    if 'nip' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/oblicz', methods=['POST'])
@login_required
def oblicz_beben():
    """Oblicz bęben (tylko dla zalogowanych)"""
    try:
        # Pobranie danych z formularza
        nazwa_kabla = request.form.get('nazwa_kabla', '').strip()
        liczba_przekroj = request.form.get('liczba_przekroj', '').strip()
        dlugosc_kabla = request.form.get('dlugosc_kabla', '')
        
        # Walidacja danych
        errors = validate_input_data(nazwa_kabla, liczba_przekroj, dlugosc_kabla)
        if errors:
            flash(f"Błędy walidacji: {'; '.join(errors)}", 'error')
            return redirect(url_for('calculator'))
        
        dlugosc_kabla = float(dlugosc_kabla)
        kable_data, bebny_data = load_excel_data()
        
        # Znajdź wybrany kabel
        wybrany_kabel = None
        for kabel in kable_data:
            if (kabel.get('Nazwa') == nazwa_kabla and 
                kabel.get('Liczba i przekrój żył') == liczba_przekroj):
                wybrany_kabel = kabel
                break
        
        if not wybrany_kabel:
            flash("Nie znaleziono kabla o podanych parametrach.", 'error')
            return redirect(url_for('calculator'))
        
        # Parametry kabla
        srednica_kabla = wybrany_kabel.get('średnica zewnętrzna kabla', 0) / 10
        promień_gięcia = wybrany_kabel.get('promień gięcia', 0) / 10
        masa_kabla_na_km = wybrany_kabel.get('Masa kg/km', 0)
        
        if dlugosc_kabla < 400:
            promień_gięcia -= 5
        
        # Znajdź wszystkie odpowiednie bębny
        odpowiednie_bebny = find_suitable_drums(
            srednica_kabla, promień_gięcia, dlugosc_kabla, bebny_data, masa_kabla_na_km
        )
        
        if not odpowiednie_bebny:
            flash("Nie znaleziono odpowiedniego bębna.", 'error')
            return redirect(url_for('calculator'))
        
        # Najlepszy bęben (pierwszy z posortowanej listy)
        najlepszy_beben = odpowiednie_bebny[0]
        
        # Przekaż dane do template
        return render_template('calculator.html',
                             wynik_data={
                                 'nazwa_kabla': nazwa_kabla,
                                 'przekroj': liczba_przekroj,
                                 'dlugosc': dlugosc_kabla,
                                 'srednica_bebna': najlepszy_beben['beben']['Średnica'],
                                 'laczna_masa': najlepszy_beben['suma_wag'],
                                 'szczegoly': najlepszy_beben
                             },
                             opcje_kabli=get_kable_options())
                             
    except Exception as e:
        logging.error(f"Błąd podczas obliczeń: {e}")
        flash(f"Wystąpił błąd podczas obliczeń: {str(e)}", 'error')
        return redirect(url_for('calculator'))

@app.route('/api/cable-options')
@login_required
def api_cable_options():
    """API endpoint dla opcji kabli (tylko dla zalogowanych)"""
    try:
        return jsonify(get_kable_options())
    except Exception as e:
        logging.error(f"Błąd API cable-options: {e}")
        return jsonify({"error": "Błąd serwera"}), 500

@app.errorhandler(404)
def not_found(error):
    return render_template('error.html', error="Strona nie została znaleziona"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', error="Wewnętrzny błąd serwera"), 500

# Inicializacja przy starcie
init_mock_data()

# Dla Vercel
if __name__ == '__main__':
    app.run(debug=False)
